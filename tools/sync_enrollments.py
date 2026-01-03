# tools/sync_enrollments.py
from __future__ import annotations

from openpyxl import load_workbook
from django.db import transaction

from quiz.models import Participant, Enrollment

PATH = "data/participants_import_ready.xlsx"

# domain -> (مجرد تسمية للمراجعة/الطباعة)
MAP = {
    "deputy": "وكيل/وكيلة",
    "principal": "مدير/مديرة",
    "supervisor": "مشرف/مشرفة",
    # عدّلها حسب الدومينات الفعلية عندك
}

SHEET_NAME = None  # اتركها None لو أول شيت
COL_NAT = "A"      # السجل المدني
COL_DOMAIN = "B"   # المجال (domain)

def norm(v) -> str:
    return str(v or "").strip()

def main():
    wb = load_workbook(PATH)
    ws = wb[SHEET_NAME] if SHEET_NAME else wb.active

    rows = ws.max_row
    print("Rows in sheet:", rows)

    created = 0
    updated = 0
    skipped_no_participant = 0
    skipped_bad_domain = 0

    with transaction.atomic():
        for r in range(2, rows + 1):  # نفترض صف 1 عناوين
            nat = norm(ws[f"{COL_NAT}{r}"].value)
            domain = norm(ws[f"{COL_DOMAIN}{r}"].value).lower()

            if not nat:
                continue

            if not domain or (MAP and domain not in MAP):
                skipped_bad_domain += 1
                continue

            participant = Participant.objects.filter(national_id=nat).first()
            if not participant:
                skipped_no_participant += 1
                continue

            obj, was_created = Enrollment.objects.get_or_create(
                participant=participant,
                domain=domain,
                defaults={"is_allowed": True},
            )
            if was_created:
                created += 1
            else:
                if not obj.is_allowed:
                    obj.is_allowed = True
                    obj.save(update_fields=["is_allowed"])
                    updated += 1

    print("Enrollments created:", created)
    print("Enrollments updated:", updated)
    print("Skipped (bad/no domain):", skipped_bad_domain)
    print("Skipped (no participant):", skipped_no_participant)
    print("Enrollments total:", Enrollment.objects.count())

if __name__ == "__main__":
    main()
