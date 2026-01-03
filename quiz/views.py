# quiz/views.py  (استبدله بالكامل)
from __future__ import annotations

import csv
import io
from dataclasses import dataclass
from datetime import datetime, timedelta
from typing import Iterable

from django.contrib import messages
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib.auth import logout as auth_logout
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import Count, Exists, OuterRef, Q
from django.http import HttpRequest, HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.views.decorators.http import require_http_methods, require_POST

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from .models import (
    Answer,
    Attempt,
    Choice,
    Enrollment,
    ExamWindow,
    Participant,
    Question,
    Quiz,
    domain_label,
)

# ======================================================
# Session Keys / Constants
# ======================================================
SESSION_PID = "participant_id"
SESSION_CONFIRMED = "participant_confirmed"
SESSION_ATTEMPT_ID = "attempt_id"

PASS_SCORE = 25  # أقل من 25 لا يرشح للمقابلة (قرار إداري)

VALID_DOMAINS = {"deputy", "counselor", "activity"}

DOMAIN_AR = {
    "deputy": "وكيل",
    "counselor": "موجّه طلابي",
    "activity": "رائد نشاط",
}

# ✅ ربط صارم بين المجال وعنوان الاختبار (لا fallback)
# مهم: لازم تكون عناوين الـ Quiz في لوحة الإدارة مطابقة لهذه النصوص حرفيًا (أو غيّرها بما عندك)
QUIZ_TITLE_BY_DOMAIN = {
    "deputy": "وكيل",
    "counselor": "موجه طلابي",
    "activity": "رائد نشاط",
}

_AR_DIGITS = str.maketrans("٠١٢٣٤٥٦٧٨٩۰۱۲۳۴۵۶۷۸۹", "01234567890123456789")


# ======================================================
# Helpers
# ======================================================
def _norm(s: str | None) -> str:
    return (s or "").strip()


def _digits(s: str | None) -> str:
    return _norm(str(s or "")).translate(_AR_DIGITS)


def _domain_ar(domain: str | None) -> str:
    d = _norm(domain).lower()
    return DOMAIN_AR.get(d) or domain_label(d) or (d or "")


def _now() -> datetime:
    return timezone.now()


def _get_active_window(domain: str) -> ExamWindow | None:
    t = _now()
    return (
        ExamWindow.objects.filter(is_active=True, domain=domain, starts_at__lte=t, ends_at__gte=t)
        .order_by("-starts_at", "-id")
        .first()
    )


def _is_in_window(domain: str) -> bool:
    return _get_active_window(domain) is not None


def _build_qs_base(request: HttpRequest, *, drop: Iterable[str] = ("page",)) -> str:
    qd = request.GET.copy()
    for k in drop:
        qd.pop(k, None)
    return qd.urlencode()


def _attempt_questions_qs(quiz: Quiz):
    return Question.objects.filter(quiz=quiz).order_by("order", "id")


def _ensure_answer_row(attempt: Attempt, q: Question) -> Answer:
    """
    ✅ تحسين بسيط للوقت:
    - إذا Answer موجود يرجع كما هو
    - إذا غير موجود ينشئ Answer ويثبت started_at (لو الحقل موجود) مرة واحدة
    """
    ans = Answer.objects.filter(attempt=attempt, question=q).first()
    if ans:
        return ans

    ans = Answer.objects.create(attempt=attempt, question=q)

    # تثبيت started_at عند الإنشاء (حتى يكون العدّاد ثابت)
    if hasattr(ans, "started_at") and getattr(ans, "started_at", None) is None:
        try:
            ans.started_at = _now()
            ans.save(update_fields=["started_at"])
        except Exception:
            pass

    return ans


def _attempt_current_question(attempt: Attempt) -> Question | None:
    if attempt.is_finished:
        return None
    qs = _attempt_questions_qs(attempt.quiz)
    idx = int(attempt.current_index or 0)
    return qs.all()[idx : idx + 1].first()


def _attempt_total_questions(attempt: Attempt) -> int:
    return _attempt_questions_qs(attempt.quiz).count()


def _answer_started_at(ans: Answer) -> datetime:
    if hasattr(ans, "started_at"):
        v = getattr(ans, "started_at")
        if v:
            return v
    if hasattr(ans, "created_at"):
        v = getattr(ans, "created_at")
        if v:
            return v
    return _now()


def _attempt_deadline(attempt: Attempt, ans: Answer) -> datetime:
    sec = int(getattr(attempt.quiz, "per_question_seconds", 50) or 50)
    return _answer_started_at(ans) + timedelta(seconds=sec)


def _attempt_remaining_seconds(attempt: Attempt, ans: Answer) -> int:
    deadline = _attempt_deadline(attempt, ans)
    return int(max(0, (deadline - _now()).total_seconds()))


def _attempt_is_overdue(attempt: Attempt, ans: Answer) -> bool:
    return (not attempt.is_finished) and (_attempt_remaining_seconds(attempt, ans) <= 0)


def _finish_attempt(attempt: Attempt, reason: str = "normal") -> None:
    if attempt.is_finished:
        return
    attempt.is_finished = True
    attempt.finished_reason = reason
    attempt.finished_at = _now()
    attempt.save(update_fields=["is_finished", "finished_reason", "finished_at"])


# ✅ اختيار Quiz “صارم” حسب المجال (بدون أي fallback)
def _quiz_title_for_domain(domain: str) -> str | None:
    d = _norm(domain).lower()
    return QUIZ_TITLE_BY_DOMAIN.get(d)


def _get_quiz_for_domain(domain: str, *, active_only: bool = True) -> Quiz | None:
    title = _quiz_title_for_domain(domain)
    if not title:
        return None
    qs = Quiz.objects.filter(title=title)
    if active_only:
        qs = qs.filter(is_active=True)
    return qs.order_by("-id").first()


# ======================================================
# Excel Import helpers (Questions)
# ======================================================
def _headers_map(ws: Worksheet) -> dict[str, int]:
    headers: dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        key = _norm(str(ws.cell(row=1, column=c).value or "")).lower()
        if key:
            headers[key] = c
    return headers


def _col(headers: dict[str, int], *names: str) -> int | None:
    for n in names:
        k = n.lower()
        if k in headers:
            return headers[k]
    return None


def _sheet_name_to_domain(sheet_name: str) -> str | None:
    s = _norm(sheet_name).lower()
    if s in VALID_DOMAINS:
        return s
    if "وكيل" in s:
        return "deputy"
    if "موجه" in s or "موجّه" in s:
        return "counselor"
    if "نشاط" in s:
        return "activity"
    return None


def _parse_correct(v: str) -> int:
    corr = _norm(v).upper()
    if corr in {"A", "B", "C", "D"}:
        return {"A": 0, "B": 1, "C": 2, "D": 3}[corr]
    try:
        n = int(_digits(corr))
        if 1 <= n <= 4:
            return n - 1
    except Exception:
        pass
    return 0


def _get_or_create_quiz_for_domain(domain: str) -> Quiz:
    """
    ✅ ينشئ/يجلب Quiz واحد “محدد” لكل مجال بعنوان ثابت من QUIZ_TITLE_BY_DOMAIN.
    - هذا هو أساس منع تداخل الأسئلة بين المجالات.
    """
    title = _quiz_title_for_domain(domain)
    if not title:
        raise ValueError("تعذر تحديد عنوان الاختبار للمجال.")

    q = Quiz.objects.filter(title=title).order_by("-id").first()
    if q:
        return q

    # عند الإنشاء: افتراضيًا غير نشط (لتفعيله يدويًا من لوحة الإدارة)
    return Quiz.objects.create(title=title, is_active=False)


def _replace_questions_for_quiz(quiz: Quiz, ws: Worksheet) -> tuple[int, int]:
    headers = _headers_map(ws)

    c_order = _col(headers, "order", "ترتيب", "رقم", "no")
    c_text = _col(headers, "text", "نص السؤال", "السؤال", "question")
    c_a = _col(headers, "a", "خيار1", "option1", "choice1")
    c_b = _col(headers, "b", "خيار2", "option2", "choice2")
    c_c = _col(headers, "c", "خيار3", "option3", "choice3")
    c_d = _col(headers, "d", "خيار4", "option4", "choice4")
    c_correct = _col(headers, "correct", "الصحيح", "answer")

    if not c_text or not c_a or not c_b or not c_correct:
        raise ValueError("الأعمدة المطلوبة: text + a + b + correct (والبقية اختياري).")

    # ✅ استبدال كامل لأسئلة هذا الـ Quiz فقط
    Question.objects.filter(quiz=quiz).delete()

    created_q = 0
    created_c = 0

    for r in range(2, ws.max_row + 1):
        q_text = _norm(ws.cell(row=r, column=c_text).value)
        if not q_text:
            continue

        try:
            order = (
                int(_digits(str(ws.cell(row=r, column=c_order).value or "")) or "1")
                if c_order
                else (created_q + 1)
            )
        except Exception:
            order = created_q + 1

        qobj = Question.objects.create(quiz=quiz, order=order, text=q_text)
        created_q += 1

        opts = [
            _norm(ws.cell(row=r, column=c_a).value) if c_a else "",
            _norm(ws.cell(row=r, column=c_b).value) if c_b else "",
            _norm(ws.cell(row=r, column=c_c).value) if c_c else "",
            _norm(ws.cell(row=r, column=c_d).value) if c_d else "",
        ]
        correct_idx = _parse_correct(str(ws.cell(row=r, column=c_correct).value or ""))

        for i, t in enumerate(opts):
            if not t:
                continue
            Choice.objects.create(question=qobj, text=t, is_correct=(i == correct_idx))
            created_c += 1

    return created_q, created_c


# ======================================================
# Public
# ======================================================
@require_http_methods(["GET"])
def home(request: HttpRequest) -> HttpResponse:
    return redirect("quiz:login")


@require_POST
def reset_session_view(request: HttpRequest) -> HttpResponse:
    for k in [SESSION_PID, SESSION_CONFIRMED, SESSION_ATTEMPT_ID, "selected_domain", "open_domains"]:
        request.session.pop(k, None)

    try:
        request.session.flush()
    except Exception:
        try:
            request.session.cycle_key()
        except Exception:
            pass

    messages.success(request, "تمت تهيئة الجلسة بنجاح.")
    return redirect("quiz:login")


@require_http_methods(["GET", "POST"])
def login_view(request: HttpRequest) -> HttpResponse:
    if request.method == "GET":
        locked_domain = ""
        locked_domain_label = ""

        pid = request.session.get(SESSION_PID)
        if pid:
            p = Participant.objects.filter(id=pid).first()
            if p:
                locked_domain = _norm(getattr(p, "locked_domain", "")).lower()
                if locked_domain:
                    locked_domain_label = _domain_ar(locked_domain)

        return render(
            request,
            "quiz/login.html",
            {
                "pass_score": PASS_SCORE,
                "locked_domain": locked_domain,
                "locked_domain_label": locked_domain_label,
                "active_domain_label": "",
            },
        )

    nid = _digits(request.POST.get("national_id"))
    last4 = _digits(request.POST.get("phone_last4"))

    if not nid:
        messages.error(request, "أدخل رقم الهوية/السجل.")
        return redirect("quiz:login")
    if not last4:
        messages.error(request, "أدخل آخر 4 أرقام من الجوال.")
        return redirect("quiz:login")

    p = Participant.objects.filter(national_id=nid).first()
    if not p:
        messages.error(request, "بيانات الدخول غير صحيحة.")
        return redirect("quiz:login")

    if _digits(getattr(p, "phone_last4", "")) and _digits(getattr(p, "phone_last4", "")) != last4:
        messages.error(request, "بيانات الدخول غير صحيحة.")
        return redirect("quiz:login")

    if not getattr(p, "is_allowed", True):
        messages.error(request, "غير مسموح لك بالدخول (قرار إداري).")
        return redirect("quiz:login")

    # ✅ إذا سبق وأنه بدأ/أنهى اختبارًا: ممنوع دخول أي اختبار آخر
    # (مسموح فقط بالعودة لنفس المحاولة لو كانت مفتوحة — هذا يتم في confirm/question)
    if getattr(p, "has_taken_exam", False):
        # لا نوجهه مباشرة للـ finish لأن قد يكون عنده محاولة مفتوحة (نخليه يمر عبر confirm)
        # لكن لو ما عنده أي محاولة مفتوحة لاحقاً سيُمنع
        pass

    allowed_domains = list(
        Enrollment.objects.filter(participant=p, is_allowed=True).values_list("domain", flat=True)
    )

    if not allowed_domains:
        messages.error(request, "لا توجد لك صلاحية لأي مجال.")
        return redirect("quiz:login")

    locked = _norm(getattr(p, "locked_domain", "")).lower()

    if locked:
        if locked not in allowed_domains:
            messages.error(request, "بيانات المجال غير متوافقة مع صلاحياتك. تواصل مع الإدارة.")
            return redirect("quiz:login")

        if not _is_in_window(locked):
            messages.error(request, f"الاختبار غير متاح الآن لمجال {_domain_ar(locked)} (خارج النافذة الزمنية).")
            return redirect("quiz:login")

        chosen_domain = locked
        request.session["open_domains"] = []
    else:
        open_now = [d for d in allowed_domains if _is_in_window(d)]

        if not open_now:
            messages.error(request, "لا يوجد اختبار متاح الآن لك (خارج النوافذ الزمنية).")
            return redirect("quiz:login")

        # ✅ إذا أكثر من مجال متاح الآن: اترك له حرية الاختيار
        if len(open_now) > 1:
            request.session[SESSION_PID] = p.id
            request.session[SESSION_CONFIRMED] = False
            request.session["selected_domain"] = ""
            request.session["open_domains"] = open_now
            return redirect("quiz:choose_domain")

        chosen_domain = open_now[0]
        request.session["open_domains"] = []

    request.session[SESSION_PID] = p.id
    request.session[SESSION_CONFIRMED] = False
    request.session["selected_domain"] = chosen_domain
    return redirect("quiz:confirm")


@require_http_methods(["GET", "POST"])
def choose_domain_view(request: HttpRequest) -> HttpResponse:
    """
    ✅ اختيار المجال إذا كان للمرشح أكثر من مجال مفتوح الآن.
    - لا يرفع has_taken_exam هنا
    - يقفل locked_domain عند الاختيار (حتى لا يغير المجال)
    """
    pid = request.session.get(SESSION_PID)
    if not pid:
        return redirect("quiz:login")

    p = get_object_or_404(Participant, id=pid)

    locked = _norm(getattr(p, "locked_domain", "")).lower()
    if locked:
        request.session["selected_domain"] = locked
        return redirect("quiz:confirm")

    open_domains = request.session.get("open_domains") or []
    open_domains = [_norm(d).lower() for d in open_domains if _norm(d).lower() in VALID_DOMAINS]

    # إذا ما عندنا بالقائمة لسبب ما: نبنيها من DB (دفاعيًا)
    if not open_domains:
        allowed_domains = list(
            Enrollment.objects.filter(participant=p, is_allowed=True).values_list("domain", flat=True)
        )
        open_domains = [d for d in allowed_domains if _is_in_window(d)]

    if not open_domains:
        messages.error(request, "لا يوجد اختبار متاح الآن لك (خارج النوافذ الزمنية).")
        return redirect("quiz:login")

    # واحد فقط -> اقفل وتابع
    if len(open_domains) == 1:
        dom = _norm(open_domains[0]).lower()
        with transaction.atomic():
            p2 = Participant.objects.select_for_update().get(id=p.id)
            if not _norm(getattr(p2, "locked_domain", "")).lower():
                p2.locked_domain = dom
                p2.locked_at = _now()
                p2.save(update_fields=["locked_domain", "locked_at"])
        request.session["selected_domain"] = dom
        return redirect("quiz:confirm")

    if request.method == "GET":
        items = [{"domain": d, "label": _domain_ar(d)} for d in open_domains]
        return render(request, "quiz/choose_domain.html", {"items": items, "p": p})

    # POST
    dom = _norm(request.POST.get("domain", "")).lower()
    if dom not in set(open_domains):
        items = [{"domain": d, "label": _domain_ar(d)} for d in open_domains]
        return render(request, "quiz/choose_domain.html", {"items": items, "p": p, "err": "اختيار غير صالح."})

    with transaction.atomic():
        p2 = Participant.objects.select_for_update().get(id=p.id)
        if not _norm(getattr(p2, "locked_domain", "")).lower():
            p2.locked_domain = dom
            p2.locked_at = _now()
            p2.save(update_fields=["locked_domain", "locked_at"])

    request.session["selected_domain"] = dom
    return redirect("quiz:confirm")


@require_POST
def logout_view(request: HttpRequest) -> HttpResponse:
    for k in [SESSION_PID, SESSION_CONFIRMED, SESSION_ATTEMPT_ID, "selected_domain", "open_domains"]:
        request.session.pop(k, None)
    messages.success(request, "تم تسجيل الخروج.")
    return redirect("quiz:login")


@require_http_methods(["GET", "POST"])
def confirm_view(request: HttpRequest) -> HttpResponse:
    pid = request.session.get(SESSION_PID)
    domain = request.session.get("selected_domain")

    if not pid:
        return redirect("quiz:login")

    p = get_object_or_404(Participant, id=pid)

    domain = _norm(domain).lower()
    if not domain:
        return redirect("quiz:choose_domain")

    if domain not in VALID_DOMAINS:
        messages.error(request, "تعذر تحديد المجال. تواصل مع الإدارة.")
        return redirect("quiz:login")

    locked = _norm(getattr(p, "locked_domain", "")).lower()
    if locked and locked != domain:
        messages.error(request, f"تم قفلك على مجال: {_domain_ar(p.locked_domain)}. لا يمكنك تغيير المجال.")
        return redirect("quiz:login")

    # ✅ اختيار صارم للاختبار حسب المجال (بدون أي fallback)
    quiz = _get_quiz_for_domain(domain, active_only=True)
    if not quiz:
        messages.error(
            request,
            f"لا يوجد اختبار نشط مضبوط لهذا المجال ({_domain_ar(domain)}). راجع عناوين الاختبارات وتفعيلها.",
        )
        return redirect("quiz:login")

    window = _get_active_window(domain)
    next_windows = (
        ExamWindow.objects.filter(is_active=True, domain=domain, starts_at__gt=_now())
        .order_by("starts_at")[:3]
    )

    if request.method == "GET":
        for w in next_windows:
            setattr(w, "domain_label", _domain_ar(getattr(w, "domain", "")))

        return render(
            request,
            "quiz/confirm.html",
            {
                "p": p,
                "participant": p,
                "domain": domain,
                "domain_label": _domain_ar(domain),
                "quiz": quiz,
                "window": window,
                "next_windows": next_windows,
            },
        )

    # =========================
    # POST: موافقة
    # =========================
    request.session[SESSION_CONFIRMED] = True

    # اقفل المجال لو لم يكن مقفولاً
    if not locked:
        p.locked_domain = domain
        p.locked_at = _now()
        p.save(update_fields=["locked_domain", "locked_at"])

    if not window:
        messages.error(request, "لا توجد نافذة اختبار نشطة الآن لهذا المجال.")
        return redirect("quiz:confirm")

    # ✅ إذا لديه محاولة مفتوحة لنفس المجال/الاختبار -> يكملها (حتى لو has_taken_exam=True)
    a = (
        Attempt.objects.filter(participant=p, domain=domain, quiz=quiz, is_finished=False)
        .order_by("-started_at", "-id")
        .first()
    )
    if a:
        request.session[SESSION_ATTEMPT_ID] = a.id
        return redirect("quiz:question")

    # ✅ لا يوجد Attempt مفتوح: هنا هو بدء الاختبار فعلياً
    # - إن كان has_taken_exam=True -> ممنوع إنشاء محاولة جديدة
    if getattr(p, "has_taken_exam", False):
        messages.error(request, "تم تنفيذ الاختبار مسبقًا ولا يمكن إنشاء محاولة جديدة.")
        return redirect("quiz:finish")

    # ✅ إنشاء Attempt جديد + رفع has_taken_exam=True مرة واحدة فقط
    with transaction.atomic():
        p2 = Participant.objects.select_for_update().get(id=p.id)
        if getattr(p2, "has_taken_exam", False):
            messages.error(request, "تم تنفيذ الاختبار مسبقًا ولا يمكن إنشاء محاولة جديدة.")
            return redirect("quiz:finish")

        a = Attempt.objects.create(
            participant=p2,
            quiz=quiz,
            domain=domain,
            session_key=_norm(getattr(request.session, "session_key", "")) or _norm(request.session.session_key),
            started_ip=request.META.get("REMOTE_ADDR"),
            user_agent=_norm(request.META.get("HTTP_USER_AGENT"))[:255] or None,
        )

        p2.has_taken_exam = True
        p2.save(update_fields=["has_taken_exam"])

    request.session[SESSION_ATTEMPT_ID] = a.id
    return redirect("quiz:question")


@require_http_methods(["GET", "POST"])
def question_view(request: HttpRequest) -> HttpResponse:
    pid = request.session.get(SESSION_PID)
    confirmed = bool(request.session.get(SESSION_CONFIRMED))
    attempt_id = request.session.get(SESSION_ATTEMPT_ID)

    if not pid or not confirmed or not attempt_id:
        return redirect("quiz:login")

    a = get_object_or_404(Attempt.objects.select_related("participant", "quiz"), id=attempt_id)

    # ✅ حماية إضافية: لا تسمح بمحاولة ليست لهذا المستخدم
    if int(a.participant_id) != int(pid):
        messages.error(request, "جلسة غير صالحة. أعد تسجيل الدخول.")
        return redirect("quiz:login")

    if a.is_finished:
        return redirect("quiz:finish")

    p = a.participant
    dom = _norm(getattr(a, "domain", "")).lower()

    locked = _norm(getattr(p, "locked_domain", "")).lower()
    if locked and locked != dom:
        messages.error(request, f"تم قفلك على مجال: {_domain_ar(p.locked_domain)}. لا يمكنك تغيير المجال.")
        _finish_attempt(a, reason="blocked")
        return redirect("quiz:finish")

    # قفل المجال إذا لم يكن مقفولًا (احتياط)
    if not locked:
        p.locked_domain = dom
        p.locked_at = _now()
        p.save(update_fields=["locked_domain", "locked_at"])

    q = _attempt_current_question(a)
    if not q:
        _finish_attempt(a, reason="normal")
        return redirect("quiz:finish")

    ans = _ensure_answer_row(a, q)

    # ✅ إذا انتهى الوقت عند GET: تقدم تلقائي للسؤال التالي
    if _attempt_is_overdue(a, ans) and request.method == "GET":
        a.timed_out_count = int(getattr(a, "timed_out_count", 0) or 0) + 1
        a.current_index = int(getattr(a, "current_index", 0) or 0) + 1
        a.save(update_fields=["timed_out_count", "current_index"])
        return redirect("quiz:question")

    if request.method == "GET":
        remaining_seconds = _attempt_remaining_seconds(a, ans)
        total = _attempt_total_questions(a)
        index = int(getattr(a, "current_index", 0) or 0) + 1
        choices = Choice.objects.filter(question=q).order_by("id")

        return render(
            request,
            "quiz/question.html",
            {
                "attempt": a,
                "question": q,
                "choices": choices,
                "index": index,
                "total": total,
                "remaining_seconds": remaining_seconds,
            },
        )

    # =========================
    # ✅ POST: دعم skip
    # =========================
    skip = request.POST.get("skip") == "1"

    choice_id = None if skip else request.POST.get("choice")
    selected = Choice.objects.filter(id=choice_id, question=q).first() if choice_id else None

    if _attempt_is_overdue(a, ans):
        a.timed_out_count = int(getattr(a, "timed_out_count", 0) or 0) + 1
        ans.answered_at = _now()
        ans.selected_choice = None
        try:
            ans.save(update_fields=["answered_at", "selected_choice"])
        except Exception:
            ans.save()
    else:
        ans.answered_at = _now()
        ans.selected_choice = None if skip else selected
        try:
            ans.save(update_fields=["answered_at", "selected_choice"])
        except Exception:
            ans.save()

        if (not skip) and selected and bool(getattr(selected, "is_correct", False)):
            a.score = int(getattr(a, "score", 0) or 0) + 1

    a.current_index = int(getattr(a, "current_index", 0) or 0) + 1

    total = _attempt_total_questions(a)
    if a.current_index >= total:
        _finish_attempt(a, reason="normal")
        a.save(update_fields=["score", "current_index"])
        return redirect("quiz:finish")

    a.save(update_fields=["score", "current_index"])
    return redirect("quiz:question")


@require_http_methods(["GET"])
def finish_view(request: HttpRequest) -> HttpResponse:
    # ✅ لا نعرض النتيجة للمختبر
    return render(request, "quiz/finish.html", {})


# ======================================================
# Staff Auth
# ======================================================
@require_POST
@staff_member_required
def staff_logout_view(request: HttpRequest) -> HttpResponse:
    auth_logout(request)
    messages.success(request, "تم تسجيل الخروج من الإدارة.")
    return redirect("/admin/login/")


# ======================================================
# Staff: Manage (Attempts + Not Tested)
# ======================================================
@dataclass
class _Filters:
    q: str = ""
    status: str = "all"
    cand: str = ""
    domain: str = ""
    quiz: str = ""
    fr: str = ""
    min_score: str = ""
    from_: str = ""
    to: str = ""
    sort: str = "-started_at"


@staff_member_required
@require_http_methods(["GET"])
def staff_manage_view(request: HttpRequest) -> HttpResponse:
    f = _Filters(
        q=_digits(request.GET.get("q")),
        status=_norm(request.GET.get("status", "all")) or "all",
        cand=_norm(request.GET.get("cand")),
        domain=_norm(request.GET.get("domain")),
        quiz=_norm(request.GET.get("quiz")),
        fr=_norm(request.GET.get("fr")),
        min_score=_digits(request.GET.get("min_score")),
        from_=_norm(request.GET.get("from")),
        to=_norm(request.GET.get("to")),
        sort=_norm(request.GET.get("sort", "-started_at")) or "-started_at",
    )

    not_tested_mode = request.GET.get("not_tested") == "1"
    quizzes = Quiz.objects.order_by("-id")

    allowed_enrollments = Enrollment.objects.filter(is_allowed=True)

    attempt_exists = Attempt.objects.filter(participant_id=OuterRef("participant_id"))
    allowed_with_attempt_flag = allowed_enrollments.annotate(has_attempt=Exists(attempt_exists))

    total_candidates = allowed_enrollments.values("participant_id").distinct().count()
    tested_candidates = allowed_with_attempt_flag.filter(has_attempt=True).values("participant_id").distinct().count()
    not_tested_candidates = allowed_with_attempt_flag.filter(has_attempt=False).values("participant_id").distinct().count()

    by_domain_attempts = Attempt.objects.values("domain").annotate(total=Count("id")).order_by("-total")
    by_domain_tested = [
        {"domain": row["domain"], "label": _domain_ar(row["domain"]), "total": row["total"]}
        for row in by_domain_attempts
    ]

    by_domain_not_tested_rows = (
        allowed_with_attempt_flag.filter(has_attempt=False)
        .values("domain")
        .annotate(total=Count("participant_id", distinct=True))
        .order_by("-total")
    )
    by_domain_not_tested = [
        {"domain": row["domain"], "label": _domain_ar(row["domain"]), "total": row["total"]}
        for row in by_domain_not_tested_rows
    ]

    by_domain_allowed_rows = (
        allowed_enrollments.values("domain")
        .annotate(total=Count("participant_id", distinct=True))
        .order_by("-total")
    )
    by_domain_allowed = [
        {"domain": row["domain"], "label": _domain_ar(row["domain"]), "total": row["total"]}
        for row in by_domain_allowed_rows
    ]

    qs_base = _build_qs_base(request, drop=("page",))
    qd_no_nt = request.GET.copy()
    qd_no_nt.pop("not_tested", None)
    qd_no_nt.pop("page", None)
    qs_base_no_nt = qd_no_nt.urlencode()

    # =========================
    # Mode: Not tested
    # =========================
    if not_tested_mode:
        qs = Enrollment.objects.filter(is_allowed=True).select_related("participant")
        qs = qs.annotate(has_attempt=Exists(attempt_exists)).filter(has_attempt=False)

        if f.q:
            qs = qs.filter(Q(participant__national_id__icontains=f.q) | Q(participant__full_name__icontains=f.q))
        if f.domain:
            qs = qs.filter(domain=f.domain)

        qs = qs.order_by("-id")

        paginator = Paginator(qs, 50)
        page_obj = paginator.get_page(request.GET.get("page") or 1)

        for e in page_obj:
            setattr(e, "display_domain", _domain_ar(e.domain))

        return render(
            request,
            "quiz/staff_manage.html",
            {
                "mode": "not_tested",
                "filters": {
                    "q": f.q,
                    "status": f.status,
                    "cand": f.cand,
                    "domain": f.domain,
                    "quiz": f.quiz,
                    "fr": f.fr,
                    "min_score": f.min_score,
                    "from": f.from_,
                    "to": f.to,
                    "sort": f.sort,
                },
                "pass_score": PASS_SCORE,
                "quizzes": quizzes,
                "not_tested": page_obj,
                "attempts": None,
                "qs_base": qs_base,
                "qs_base_no_nt": qs_base_no_nt,
                "kpi_people": {
                    "total_candidates": total_candidates,
                    "tested_candidates": tested_candidates,
                    "not_tested_candidates": not_tested_candidates,
                },
                "by_domain_tested": by_domain_tested,
                "by_domain_not_tested": by_domain_not_tested,
                "by_domain_allowed": by_domain_allowed,
            },
        )

    # =========================
    # Mode: Attempts
    # =========================
    qs = Attempt.objects.select_related("participant", "quiz")

    if f.q:
        qs = qs.filter(Q(participant__national_id__icontains=f.q) | Q(participant__full_name__icontains=f.q))
    if f.domain:
        qs = qs.filter(domain=f.domain)
    if f.quiz:
        qs = qs.filter(quiz_id=f.quiz)

    if f.status == "running":
        qs = qs.filter(is_finished=False)
    elif f.status == "finished":
        qs = qs.filter(is_finished=True)

    if f.fr:
        qs = qs.filter(finished_reason=f.fr)

    if f.min_score:
        try:
            qs = qs.filter(score__gte=int(f.min_score))
        except Exception:
            pass

    if f.from_:
        try:
            dt = datetime.fromisoformat(f.from_)
            qs = qs.filter(started_at__date__gte=dt.date())
        except Exception:
            pass

    if f.to:
        try:
            dt = datetime.fromisoformat(f.to)
            qs = qs.filter(started_at__date__lte=dt.date())
        except Exception:
            pass

    if f.cand == "yes":
        qs = qs.filter(is_finished=True, score__gte=PASS_SCORE)
    elif f.cand == "no":
        qs = qs.filter(is_finished=True, score__lt=PASS_SCORE)

    allowed_sorts = {"-started_at", "started_at", "-score", "score"}
    if f.sort not in allowed_sorts:
        f.sort = "-started_at"
    qs = qs.order_by(f.sort, "-id")

    paginator = Paginator(qs, 50)
    page_obj = paginator.get_page(request.GET.get("page") or 1)

    for a in page_obj:
        setattr(a, "display_score", int(getattr(a, "score", 0) or 0))
        setattr(a, "display_domain", _domain_ar(a.domain))

    return render(
        request,
        "quiz/staff_manage.html",
        {
            "mode": "attempts",
            "filters": {
                "q": f.q,
                "status": f.status,
                "cand": f.cand,
                "domain": f.domain,
                "quiz": f.quiz,
                "fr": f.fr,
                "min_score": f.min_score,
                "from": f.from_,
                "to": f.to,
                "sort": f.sort,
            },
            "pass_score": PASS_SCORE,
            "quizzes": quizzes,
            "attempts": page_obj,
            "not_tested": None,
            "qs_base": qs_base,
            "qs_base_no_nt": qs_base_no_nt,
            "kpi_people": {
                "total_candidates": total_candidates,
                "tested_candidates": tested_candidates,
                "not_tested_candidates": not_tested_candidates,
            },
            "by_domain_tested": by_domain_tested,
            "by_domain_not_tested": by_domain_not_tested,
            "by_domain_allowed": by_domain_allowed,
        },
    )


# ======================================================
# Staff: Attempt detail + actions
# ======================================================
@staff_member_required
@require_http_methods(["GET"])
def staff_attempt_detail_view(request: HttpRequest, attempt_id: int) -> HttpResponse:
    a = get_object_or_404(Attempt.objects.select_related("participant", "quiz"), id=attempt_id)

    total = _attempt_questions_qs(a.quiz).count()
    answers = (
        Answer.objects.filter(attempt=a)
        .select_related("question", "selected_choice")
        .order_by("question__order", "id")
    )

    answered = answers.filter(answered_at__isnull=False).count()
    progress_pct = int(round((answered / total) * 100)) if total else 0

    return render(
        request,
        "quiz/staff_attempt_detail.html",
        {
            "a": a,
            "attempt": a,
            "answers": answers,
            "total": total,
            "answered": answered,
            "progress_pct": progress_pct,
            "pass_score": PASS_SCORE,
        },
    )


@staff_member_required
@require_http_methods(["GET"])
def staff_attempt_finish_confirm_view(request: HttpRequest, attempt_id: int) -> HttpResponse:
    a = get_object_or_404(Attempt.objects.select_related("participant", "quiz"), id=attempt_id)
    return render(request, "quiz/staff_attempt_finish_confirm.html", {"a": a})


@staff_member_required
@require_http_methods(["GET"])
def staff_attempt_reset_confirm_view(request: HttpRequest, attempt_id: int) -> HttpResponse:
    a = get_object_or_404(Attempt.objects.select_related("participant", "quiz"), id=attempt_id)
    return render(request, "quiz/staff_attempt_reset_confirm.html", {"a": a})


@staff_member_required
@require_POST
def staff_force_finish_attempt_view(request: HttpRequest, attempt_id: int) -> HttpResponse:
    a = get_object_or_404(Attempt, id=attempt_id)
    if not a.is_finished:
        a.is_finished = True
        a.finished_reason = "forced"
        a.finished_at = _now()
        a.save(update_fields=["is_finished", "finished_reason", "finished_at"])
        messages.success(request, "تم إنهاء المحاولة إدارياً.")
    return redirect("quiz:staff_attempt_detail", attempt_id=a.id)


@staff_member_required
@require_POST
def staff_reset_attempt_view(request: HttpRequest, attempt_id: int) -> HttpResponse:
    a = get_object_or_404(Attempt.objects.select_related("participant"), id=attempt_id)

    with transaction.atomic():
        # حذف الإجابات
        Answer.objects.filter(attempt=a).delete()

        # إعادة المحاولة
        a.current_index = 0
        a.score = 0
        a.is_finished = False
        a.finished_reason = "normal"
        a.finished_at = None
        a.timed_out_count = 0
        a.save(
            update_fields=[
                "current_index",
                "score",
                "is_finished",
                "finished_reason",
                "finished_at",
                "timed_out_count",
            ]
        )

        # ✅ إعادة المرشح لحالة قبل الاختبار (للسماح بإعادة المحاولة فعلياً)
        p = a.participant
        p.has_taken_exam = False
        p.locked_domain = ""
        p.locked_at = None
        p.save(update_fields=["has_taken_exam", "locked_domain", "locked_at"])

    messages.success(request, "تمت إعادة فتح المحاولة بنجاح (مع إعادة تهيئة حالة المرشح).")
    return redirect("quiz:staff_attempt_detail", attempt_id=a.id)


# ======================================================
# Staff: Imports
# ======================================================
@staff_member_required
@require_http_methods(["GET", "POST"])
def staff_import_participants_view(request: HttpRequest) -> HttpResponse:
    if request.method == "GET":
        return render(request, "quiz/staff_import_participants.html", {})

    up = request.FILES.get("file")
    if not up:
        messages.error(request, "ارفع ملف XLSX.")
        return redirect("quiz:staff_import_participants")

    wb = load_workbook(up)
    ws = wb.active

    headers = _headers_map(ws)

    c_nid = _col(headers, "national_id", "رقم الهوية", "الهوية", "السجل")
    c_name = _col(headers, "full_name", "الاسم", "الاسم الكامل")
    c_last4 = _col(headers, "phone_last4", "آخر4", "اخر4", "آخر 4", "اخر 4")
    c_domain = _col(headers, "domain", "المجال")
    c_allowed = _col(headers, "is_allowed", "allowed", "مسموح", "is_allowed_domain")

    if not c_nid or not c_name or not c_domain:
        messages.error(request, "الأعمدة المطلوبة: الهوية + الاسم + المجال (domain).")
        return redirect("quiz:staff_import_participants")

    created_p = 0
    updated_p = 0
    enroll_created = 0
    enroll_updated = 0

    with transaction.atomic():
        for r in range(2, ws.max_row + 1):
            nid = _digits(ws.cell(row=r, column=c_nid).value)
            name = _norm(ws.cell(row=r, column=c_name).value)
            dom = _norm(ws.cell(row=r, column=c_domain).value).lower()
            last4 = _digits(ws.cell(row=r, column=c_last4).value) if c_last4 else ""

            if not nid or not name or dom not in VALID_DOMAINS:
                continue

            p, is_new = Participant.objects.get_or_create(
                national_id=nid,
                defaults={"full_name": name, "phone_last4": last4},
            )
            if is_new:
                created_p += 1
            else:
                changed = False
                if name and p.full_name != name:
                    p.full_name = name
                    changed = True
                if last4 and p.phone_last4 != last4:
                    p.phone_last4 = last4
                    changed = True
                if changed:
                    p.save(update_fields=["full_name", "phone_last4"])
                    updated_p += 1

            dom_allowed = True
            if c_allowed:
                v = _norm(ws.cell(row=r, column=c_allowed).value).lower()
                if v in {"0", "false", "no", "غير", "n"}:
                    dom_allowed = False

            e, e_new = Enrollment.objects.get_or_create(
                participant=p,
                domain=dom,
                defaults={"is_allowed": dom_allowed},
            )
            if e_new:
                enroll_created += 1
            else:
                if e.is_allowed != dom_allowed:
                    e.is_allowed = dom_allowed
                    e.save(update_fields=["is_allowed"])
                    enroll_updated += 1

    messages.success(
        request,
        f"تم الاستيراد: مشاركين جدد {created_p} | تحديث {updated_p} | تسجيلات جديدة {enroll_created} | تحديث {enroll_updated}",
    )
    return redirect("quiz:staff_manage")


@staff_member_required
@require_http_methods(["GET", "POST"])
def staff_import_questions_view(request: HttpRequest) -> HttpResponse:
    if request.method == "GET":
        return render(request, "quiz/staff_import_questions.html", {})

    up = request.FILES.get("file")
    if not up:
        messages.error(request, "ارفع ملف XLSX.")
        return redirect("quiz:staff_import_questions")

    preview_only = request.POST.get("preview") in {"1", "on", "true", "yes"}

    # ✅ جديد: تفعيل الاختبار تلقائياً بعد الرفع (اختياري)
    activate_after_import = request.POST.get("activate") in {"1", "on", "true", "yes"}

    wb = load_workbook(up)

    sheets: list[tuple[str, Worksheet]] = []
    for name in wb.sheetnames:
        dom = _sheet_name_to_domain(name)
        if dom:
            sheets.append((dom, wb[name]))

    if not sheets:
        messages.error(
            request,
            "لم يتم العثور على شيتات صالحة. استخدم أسماء الشيتات: deputy / counselor / activity (أو وكيل/موجه/نشاط).",
        )
        return redirect("quiz:staff_import_questions")

    summary: list[str] = []

    try:
        with transaction.atomic():
            for dom, ws in sheets:
                quiz = _get_or_create_quiz_for_domain(dom)

                if preview_only:
                    headers = _headers_map(ws)
                    c_text = _col(headers, "text", "نص السؤال", "السؤال", "question")
                    c_a = _col(headers, "a", "خيار1", "option1", "choice1")
                    c_b = _col(headers, "b", "خيار2", "option2", "choice2")
                    c_correct = _col(headers, "correct", "الصحيح", "answer")
                    if not c_text or not c_a or not c_b or not c_correct:
                        raise ValueError(f"الشيت ({ws.title}) ينقصه الأعمدة المطلوبة: text + a + b + correct.")
                    count_q = 0
                    for r in range(2, ws.max_row + 1):
                        if _norm(ws.cell(row=r, column=c_text).value):
                            count_q += 1
                    summary.append(f"{_domain_ar(dom)}: معاينة فقط — {count_q} سؤال")
                    continue

                created_q, created_c = _replace_questions_for_quiz(quiz, ws)

                # ✅ تفعيل تلقائي (لو اخترته)
                if activate_after_import and not quiz.is_active:
                    quiz.is_active = True
                    quiz.save(update_fields=["is_active"])

                summary.append(
                    f"{_domain_ar(dom)}: تم الاستبدال — {created_q} سؤال | {created_c} خيار"
                    + (" | ✅ تم التفعيل" if activate_after_import else "")
                )

    except Exception as e:
        messages.error(request, f"تعذر الاستيراد: {e}")
        return redirect("quiz:staff_import_questions")

    if preview_only:
        messages.success(request, " | ".join(summary) + " ✅")
    else:
        messages.success(request, "تم استيراد الأسئلة بنجاح ✅ " + " | ".join(summary))

    return redirect("quiz:staff_manage")


# ======================================================
# Staff: Exports (Attempts + Not Tested)
# ======================================================
def _apply_attempt_filters(request: HttpRequest, qs):
    q = _digits(request.GET.get("q"))
    status = _norm(request.GET.get("status", "all")) or "all"
    cand = _norm(request.GET.get("cand"))
    domain = _norm(request.GET.get("domain"))
    quiz = _norm(request.GET.get("quiz"))
    fr = _norm(request.GET.get("fr"))
    min_score = _digits(request.GET.get("min_score"))
    dfrom = _norm(request.GET.get("from"))
    dto = _norm(request.GET.get("to"))
    sort = _norm(request.GET.get("sort", "-started_at")) or "-started_at"

    if q:
        qs = qs.filter(Q(participant__national_id__icontains=q) | Q(participant__full_name__icontains=q))
    if domain:
        qs = qs.filter(domain=domain)
    if quiz:
        qs = qs.filter(quiz_id=quiz)

    if status == "running":
        qs = qs.filter(is_finished=False)
    elif status == "finished":
        qs = qs.filter(is_finished=True)

    if fr:
        qs = qs.filter(finished_reason=fr)

    if min_score:
        try:
            qs = qs.filter(score__gte=int(min_score))
        except Exception:
            pass

    if dfrom:
        try:
            dt = datetime.fromisoformat(dfrom)
            qs = qs.filter(started_at__date__gte=dt.date())
        except Exception:
            pass

    if dto:
        try:
            dt = datetime.fromisoformat(dto)
            qs = qs.filter(started_at__date__lte=dt.date())
        except Exception:
            pass

    if cand == "yes":
        qs = qs.filter(is_finished=True, score__gte=PASS_SCORE)
    elif cand == "no":
        qs = qs.filter(is_finished=True, score__lt=PASS_SCORE)

    allowed_sorts = {"-started_at", "started_at", "-score", "score"}
    if sort not in allowed_sorts:
        sort = "-started_at"
    return qs.order_by(sort, "-id")


@staff_member_required
@require_http_methods(["GET"])
def staff_export_csv_view(request: HttpRequest) -> HttpResponse:
    qs = _apply_attempt_filters(request, Attempt.objects.select_related("participant", "quiz"))

    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(
        [
            "attempt_id",
            "full_name",
            "national_id",
            "domain_ar",
            "quiz",
            "is_finished",
            "score",
            "candidate",
            "finished_reason",
            "timed_out_count",
            "started_at",
            "finished_at",
        ]
    )

    for a in qs.iterator():
        cand = "yes" if (a.is_finished and int(a.score or 0) >= PASS_SCORE) else "no"
        w.writerow(
            [
                a.id,
                a.participant.full_name,
                a.participant.national_id,
                _domain_ar(a.domain),
                a.quiz.title,
                int(bool(a.is_finished)),
                int(a.score or 0),
                cand,
                a.finished_reason,
                int(a.timed_out_count or 0),
                a.started_at.isoformat(sep=" ", timespec="minutes") if a.started_at else "",
                a.finished_at.isoformat(sep=" ", timespec="minutes") if a.finished_at else "",
            ]
        )

    resp = HttpResponse(buf.getvalue(), content_type="text/csv; charset=utf-8")
    resp["Content-Disposition"] = 'attachment; filename="attempts.csv"'
    return resp


@staff_member_required
@require_http_methods(["GET"])
def staff_export_xlsx_view(request: HttpRequest) -> HttpResponse:
    qs = _apply_attempt_filters(request, Attempt.objects.select_related("participant", "quiz"))

    wb = Workbook()
    ws = wb.active
    ws.title = "Attempts"

    ws.append(
        [
            "attempt_id",
            "الاسم",
            "الهوية",
            "المجال",
            "الاختبار",
            "منتهي؟",
            "الدرجة",
            f"مرشح؟ (>= {PASS_SCORE})",
            "سبب الإنهاء",
            "TO",
            "البدء",
            "الانتهاء",
        ]
    )

    for a in qs.iterator():
        ws.append(
            [
                a.id,
                a.participant.full_name,
                a.participant.national_id,
                _domain_ar(a.domain),
                a.quiz.title,
                "نعم" if a.is_finished else "لا",
                int(a.score or 0),
                "نعم" if (a.is_finished and int(a.score or 0) >= PASS_SCORE) else "لا",
                a.finished_reason,
                int(a.timed_out_count or 0),
                a.started_at.strftime("%Y-%m-%d %H:%M") if a.started_at else "",
                a.finished_at.strftime("%Y-%m-%d %H:%M") if a.finished_at else "",
            ]
        )

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)

    resp = HttpResponse(
        out.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = 'attachment; filename="attempts.xlsx"'
    return resp


@staff_member_required
@require_http_methods(["GET"])
def staff_export_not_tested_xlsx_view(request: HttpRequest) -> HttpResponse:
    q = _digits(request.GET.get("q"))
    domain = _norm(request.GET.get("domain")).lower()

    attempt_exists = Attempt.objects.filter(participant_id=OuterRef("participant_id"))

    qs = Enrollment.objects.filter(is_allowed=True).select_related("participant")
    qs = qs.annotate(has_attempt=Exists(attempt_exists)).filter(has_attempt=False)

    if q:
        qs = qs.filter(Q(participant__national_id__icontains=q) | Q(participant__full_name__icontains=q))
    if domain:
        qs = qs.filter(domain=domain)

    qs = qs.order_by("-id")

    wb = Workbook()
    ws = wb.active
    ws.title = "Not Tested"
    ws.append(["#", "الاسم", "الهوية", "المجال", "الحالة"])

    i = 0
    for e in qs.iterator():
        i += 1
        ws.append([i, e.participant.full_name, e.participant.national_id, _domain_ar(e.domain), "لم يدخل الاختبار"])

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)

    resp = HttpResponse(
        out.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = 'attachment; filename="not_tested.xlsx"'
    return resp
