from __future__ import annotations

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from quiz.models import Quiz, Question, Choice, domain_label


VALID_DOMAINS = {"deputy", "counselor", "activity"}

# ✅ ربط صارم بين المجال وعنوان الاختبار
QUIZ_TITLE_BY_DOMAIN = {
    "deputy": "وكيل",
    "counselor": "موجه طلابي",
    "activity": "رائد نشاط",
}

_AR_DIGITS = str.maketrans("٠١٢٣٤٥٦٧٨٩۰۱۲۳۴۵۶۷۸۹", "01234567890123456789")


# =========================
# Helpers
# =========================
def _norm(s) -> str:
    """Safe normalize for Excel values (str/int/float/None)."""
    if s is None:
        return ""
    if isinstance(s, (int, float)):
        try:
            if float(s).is_integer():
                return str(int(s))
        except Exception:
            pass
        return str(s)
    return str(s).strip()


def _digits(s) -> str:
    return _norm(s).translate(_AR_DIGITS)


def _domain_ar(domain: str) -> str:
    d = _norm(domain).lower()
    return domain_label(d) or d


def _headers_map(ws: Worksheet) -> dict[str, int]:
    headers: dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        key = _norm(ws.cell(row=1, column=c).value).lower()
        if key:
            headers[key] = c
    return headers


def _col(headers: dict[str, int], *names: str) -> int | None:
    for n in names:
        k = n.lower()
        if k in headers:
            return headers[k]
    return None


def _sheet_name_to_domain(sheet_name: str) -> str | None:
    s = _norm(sheet_name).lower()
    if s in VALID_DOMAINS:
        return s
    if "وكيل" in s:
        return "deputy"
    if "موجه" in s or "موجّه" in s:
        return "counselor"
    if "نشاط" in s:
        return "activity"
    return None


def _parse_correct(v) -> int:
    corr = _norm(v).upper()
    if corr in {"A", "B", "C", "D"}:
        return {"A": 0, "B": 1, "C": 2, "D": 3}[corr]
    try:
        n = int(_digits(corr))
        if 1 <= n <= 4:
            return n - 1
    except Exception:
        pass
    return 0


def _get_or_create_quiz_for_domain(domain: str) -> Quiz:
    title = QUIZ_TITLE_BY_DOMAIN.get(_norm(domain).lower())
    if not title:
        raise ValueError("تعذر تحديد عنوان الاختبار للمجال.")

    q = Quiz.objects.filter(title=title).order_by("-id").first()
    if q:
        return q

    # افتراضيًا غير نشط (إلا إذا استخدمت --activate)
    return Quiz.objects.create(title=title, is_active=False)


def _replace_questions_for_quiz(quiz: Quiz, ws: Worksheet) -> tuple[int, int]:
    headers = _headers_map(ws)

    c_order = _col(headers, "order", "ترتيب", "رقم", "no")
    c_text = _col(headers, "text", "نص السؤال", "السؤال", "question")
    c_a = _col(headers, "a", "خيار1", "option1", "choice1")
    c_b = _col(headers, "b", "خيار2", "option2", "choice2")
    c_c = _col(headers, "c", "خيار3", "option3", "choice3")
    c_d = _col(headers, "d", "خيار4", "option4", "choice4")
    c_correct = _col(headers, "correct", "الصحيح", "answer")

    if not c_text or not c_a or not c_b or not c_correct:
        raise ValueError("الأعمدة المطلوبة: text + a + b + correct (والبقية اختياري).")

    # ✅ استبدال كامل
    Question.objects.filter(quiz=quiz).delete()

    created_q = 0
    created_c = 0

    for r in range(2, ws.max_row + 1):
        q_text = _norm(ws.cell(row=r, column=c_text).value)
        if not q_text:
            continue

        # order اختياري
        try:
            order = int(_digits(ws.cell(row=r, column=c_order).value)) if c_order else (created_q + 1)
        except Exception:
            order = created_q + 1

        qobj = Question.objects.create(quiz=quiz, order=order, text=q_text)
        created_q += 1

        opts = [
            _norm(ws.cell(row=r, column=c_a).value) if c_a else "",
            _norm(ws.cell(row=r, column=c_b).value) if c_b else "",
            _norm(ws.cell(row=r, column=c_c).value) if c_c else "",
            _norm(ws.cell(row=r, column=c_d).value) if c_d else "",
        ]
        correct_idx = _parse_correct(ws.cell(row=r, column=c_correct).value)

        for i, t in enumerate(opts):
            if not t:
                continue
            Choice.objects.create(question=qobj, text=t, is_correct=(i == correct_idx))
            created_c += 1

    return created_q, created_c


# =========================
# Command
# =========================
class Command(BaseCommand):
    help = "Import/replace questions from an XLSX file on server (by sheets: deputy/counselor/activity)."

    def add_arguments(self, parser):
        parser.add_argument("path", type=str, help="Path to XLSX file, e.g. data/questions_import_ready.xlsx")
        parser.add_argument("--preview", action="store_true", help="Preview only (no delete/insert).")
        parser.add_argument("--activate", action="store_true", help="Activate quizzes after import (is_active=True).")

    def handle(self, *args, **opts):
        path = opts["path"]
        preview_only = bool(opts["preview"])
        activate = bool(opts["activate"])

        try:
            wb = load_workbook(path)
        except Exception as e:
            raise CommandError(f"تعذر فتح الملف: {path} ({e})")

        sheets: list[tuple[str, Worksheet]] = []
        for name in wb.sheetnames:
            dom = _sheet_name_to_domain(name)
            if dom:
                sheets.append((dom, wb[name]))

        if not sheets:
            raise CommandError(
                "لم يتم العثور على شيتات صالحة. استخدم أسماء الشيتات: deputy/counselor/activity (أو وكيل/موجه/نشاط)."
            )

        summary: list[str] = []

        try:
            with transaction.atomic():
                for dom, ws in sheets:
                    quiz = _get_or_create_quiz_for_domain(dom)

                    if preview_only:
                        headers = _headers_map(ws)
                        c_text = _col(headers, "text", "نص السؤال", "السؤال", "question")
                        c_a = _col(headers, "a", "خيار1", "option1", "choice1")
                        c_b = _col(headers, "b", "خيار2", "option2", "choice2")
                        c_correct = _col(headers, "correct", "الصحيح", "answer")
                        if not c_text or not c_a or not c_b or not c_correct:
                            raise ValueError(f"الشيت ({ws.title}) ينقصه الأعمدة المطلوبة: text + a + b + correct.")

                        count_q = 0
                        for r in range(2, ws.max_row + 1):
                            if _norm(ws.cell(row=r, column=c_text).value):
                                count_q += 1

                        summary.append(f"{_domain_ar(dom)}: معاينة فقط — {count_q} سؤال")
                        continue

                    created_q, created_c = _replace_questions_for_quiz(quiz, ws)

                    if activate:
                        quiz.is_active = True
                        quiz.save(update_fields=["is_active"])

                    summary.append(f"{_domain_ar(dom)}: تم الاستبدال — {created_q} سؤال | {created_c} خيار")

        except Exception as e:
            raise CommandError(f"تعذر الاستيراد: {e}")

        self.stdout.write(self.style.SUCCESS(" | ".join(summary)))
