from __future__ import annotations

from django.core.management.base import BaseCommand
from django.db import transaction
from openpyxl import load_workbook

from quiz.models import Participant, Enrollment


def norm(v) -> str:
    return str(v or "").strip()


class Command(BaseCommand):
    help = "Sync enrollments from Excel file (participant national_id + domain)."

    def add_arguments(self, parser):
        parser.add_argument(
            "--path",
            default="data/participants_import_ready.xlsx",
            help="Path to Excel file",
        )
        parser.add_argument(
            "--sheet",
            default=None,
            help="Sheet name (default: active sheet)",
        )
        parser.add_argument(
            "--col-nat",
            default="A",
            help="Column letter for national id (default: A)",
        )
        parser.add_argument(
            "--col-domain",
            default="B",
            help="Column letter for domain (default: B)",
        )
        parser.add_argument(
            "--strict-domains",
            action="store_true",
            help="If set, only allow domains listed in MAP (after normalization).",
        )

    def handle(self, *args, **opts):
        path: str = opts["path"]
        sheet_name: str | None = opts["sheet"]
        col_nat: str = opts["col_nat"]
        col_domain: str = opts["col_domain"]
        strict: bool = opts["strict_domains"]

        # ✅ domains الموجودة فعليًا في ملفك:
        # counselor / deputy / activity
        # التسمية هنا للعرض/التقارير فقط (الدومين المخزن يبقى كود ثابت).
        MAP = {
            "deputy": "وكيل/وكيلة",
            "counselor": "موجه/موجهة",
            "activity": "رائد/رائدة نشاط",
        }

        wb = load_workbook(path)
        ws = wb[sheet_name] if sheet_name else wb.active

        rows = ws.max_row
        self.stdout.write(self.style.NOTICE(f"Rows in sheet: {rows}"))

        created = 0
        updated = 0
        skipped_no_participant = 0
        skipped_bad_domain = 0

        max_len = Enrollment._meta.get_field("domain").max_length or 30

        with transaction.atomic():
            for r in range(2, rows + 1):  # row 1 headers
                nat = norm(ws[f"{col_nat}{r}"].value)
                raw_domain = norm(ws[f"{col_domain}{r}"].value)
                domain = raw_domain.lower().strip()

                if not nat:
                    continue

                # Normalization (لو جاء أي تنويعات مستقبلًا)
                ALIASES = {
                    "counselor": "counselor",
                    "deputy": "deputy",
                    "activity": "activity",
                }
                domain = ALIASES.get(domain, domain)

                # Postgres safety
                if not domain or len(domain) > max_len:
                    skipped_bad_domain += 1
                    continue

                # Strict mode: only allowed domains
                if strict and domain not in MAP:
                    skipped_bad_domain += 1
                    continue

                participant = Participant.objects.filter(national_id=nat).first()
                if not participant:
                    skipped_no_participant += 1
                    continue

                obj, was_created = Enrollment.objects.get_or_create(
                    participant=participant,
                    domain=domain,
                    defaults={"is_allowed": True},
                )
                if was_created:
                    created += 1
                else:
                    # ensure allowed stays True
                    if not obj.is_allowed:
                        obj.is_allowed = True
                        obj.save(update_fields=["is_allowed"])
                        updated += 1

        self.stdout.write(self.style.SUCCESS(f"Enrollments created: {created}"))
        self.stdout.write(self.style.SUCCESS(f"Enrollments updated: {updated}"))
        self.stdout.write(self.style.WARNING(f"Skipped (bad/no domain): {skipped_bad_domain}"))
        self.stdout.write(self.style.WARNING(f"Skipped (no participant): {skipped_no_participant}"))
        self.stdout.write(self.style.NOTICE(f"Enrollments total: {Enrollment.objects.count()}"))
