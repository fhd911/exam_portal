from __future__ import annotations

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from openpyxl import load_workbook

from quiz.models import Participant, Enrollment


VALID_DOMAINS = {"deputy", "counselor", "activity"}
_AR_DIGITS = str.maketrans("٠١٢٣٤٥٦٧٨٩۰۱۲۳۴۵۶۷۸۹", "01234567890123456789")


def _norm(s) -> str:
    if s is None:
        return ""
    if isinstance(s, (int, float)):
        try:
            if float(s).is_integer():
                return str(int(s))
        except Exception:
            pass
        return str(s)
    return str(s).strip()


def _digits(s) -> str:
    return _norm(s).translate(_AR_DIGITS)


def _headers_map(ws) -> dict[str, int]:
    headers: dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        key = _norm(ws.cell(row=1, column=c).value).lower()
        if key:
            headers[key] = c
    return headers


def _col(headers: dict[str, int], *names: str) -> int | None:
    for n in names:
        k = n.lower()
        if k in headers:
            return headers[k]
    return None


class Command(BaseCommand):
    help = "Import participants + enrollments from an XLSX file on server."

    def add_arguments(self, parser):
        parser.add_argument("path", type=str, help="Path to XLSX file, e.g. data/participants_import_ready.xlsx")

    def handle(self, *args, **opts):
        path = opts["path"]

        try:
            wb = load_workbook(path)
        except Exception as e:
            raise CommandError(f"تعذر فتح الملف: {path} ({e})")

        ws = wb.active
        headers = _headers_map(ws)

        c_nid = _col(headers, "national_id", "رقم الهوية", "الهوية", "السجل")
        c_name = _col(headers, "full_name", "الاسم", "الاسم الكامل")
        c_last4 = _col(headers, "phone_last4", "آخر4", "اخر4", "آخر 4", "اخر 4")
        c_domain = _col(headers, "domain", "المجال")
        c_allowed = _col(headers, "is_allowed", "allowed", "مسموح", "is_allowed_domain")

        if not c_nid or not c_name or not c_domain:
            raise CommandError("الأعمدة المطلوبة: الهوية + الاسم + المجال (domain).")

        created_p = 0
        updated_p = 0
        enroll_created = 0
        enroll_updated = 0

        with transaction.atomic():
            for r in range(2, ws.max_row + 1):
                nid = _digits(ws.cell(row=r, column=c_nid).value)
                name = _norm(ws.cell(row=r, column=c_name).value)
                dom = _norm(ws.cell(row=r, column=c_domain).value).lower()
                last4 = _digits(ws.cell(row=r, column=c_last4).value) if c_last4 else ""

                if not nid or not name or dom not in VALID_DOMAINS:
                    continue

                p, is_new = Participant.objects.get_or_create(
                    national_id=nid,
                    defaults={"full_name": name, "phone_last4": last4},
                )
                if is_new:
                    created_p += 1
                else:
                    changed = False
                    if name and p.full_name != name:
                        p.full_name = name
                        changed = True
                    if last4 and p.phone_last4 != last4:
                        p.phone_last4 = last4
                        changed = True
                    if changed:
                        p.save(update_fields=["full_name", "phone_last4"])
                        updated_p += 1

                dom_allowed = True
                if c_allowed:
                    v = _norm(ws.cell(row=r, column=c_allowed).value).lower()
                    if v in {"0", "false", "no", "غير", "n"}:
                        dom_allowed = False

                e, e_new = Enrollment.objects.get_or_create(
                    participant=p,
                    domain=dom,
                    defaults={"is_allowed": dom_allowed},
                )
                if e_new:
                    enroll_created += 1
                else:
                    if e.is_allowed != dom_allowed:
                        e.is_allowed = dom_allowed
                        e.save(update_fields=["is_allowed"])
                        enroll_updated += 1

        self.stdout.write(
            self.style.SUCCESS(
                f"تم الاستيراد: مشاركين جدد {created_p} | تحديث {updated_p} | تسجيلات جديدة {enroll_created} | تحديث {enroll_updated}"
            )
        )
